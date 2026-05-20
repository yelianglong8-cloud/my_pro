'''
Filename: GKAN_Pattern_Recognition.py
Description: Pattern recognition simulation test.
Usage: For academic research purposes only.
Note: Import device measurement data independently (memristor conductance; peak current data of GMCs;
    sigma1, sigma2, CV_amp, CV_sigma1, CV_sigma2, k and q)
    When switching device data,it is necessary to adjust the hyperparameters
    (lr, weight_decay, coefficient k, grid size, grid range and etc.) accordingly.
Characterization of Gaussian-like Basis Functions:
                                    f(x) = c * exp{-[(x-center_i)/(sigma/k)]^2}
    c: Weight coefficient of a basis function based on differential pair mechanism;
    x: Network input, represented by gate voltage;
    center_i: Center of the i-th basis function;
    sigma: When x ≤ center_i, sigma = sigma1; when x > center_i, sigma = sigma2;
    k: Factor controlling the scaling of the basis function width.
'''

import torch
import torch.nn.functional as F
import math
import pandas as pd
import numpy as np
import os
import time
import torch.nn as nn
import torch.optim as optim
import torchvision
import torchvision.transforms as transforms
from torch.utils.data import DataLoader
from tqdm import tqdm
from openpyxl import load_workbook

'''Please import the actually measured device (memristors & GMCs) data manually.'''
# Load memristor conductance data
df = pd.read_excel(r'E:\download\GMCSimu-main\GMCSimu-main\memristor conductance.xlsx', sheet_name='sheet_name', usecols='A', header=None)
memData = df.to_numpy().flatten()
memData = (memData - min(memData)) / (max(memData) - min(memData))
memDiffMat = np.subtract.outer(memData, memData)
mem_synapse = np.unique(memDiffMat.flatten())
synapse = torch.tensor(mem_synapse)
print('synapse: ',synapse.shape[0])
# Load GMC peak current data

df = pd.read_excel(r'E:\download\GMCSimu-main\GMCSimu-main\GMC peak current.xlsx', sheet_name='sheet_name', usecols='A', header=None)
aatData = df.to_numpy().flatten()
aatData = (aatData - min(aatData)) / (max(aatData) - min(aatData))
aatDiffMat = np.subtract.outer(aatData, aatData)
c_data = np.unique(aatDiffMat.flatten())
spline_coef = torch.tensor(c_data)

def changeMEM(model, synapse):
    synapse = synapse.to(next(model.parameters()).device)
    for name, param in model.named_parameters():
        if not 'spline_weight' in name:
            tensor = param.data
            diff = torch.abs(tensor.reshape(-1, 1) - synapse.reshape(1, -1))
            index = torch.argmin(diff, dim=1)
            new_tensor = synapse[index].reshape(tensor.shape)
            param.data.copy_(new_tensor)

#为一个“差分权重矩阵”建立一个快速的反向查找表
def _build_diff_lookup(aatDiffMat: np.ndarray):
    vals = aatDiffMat.reshape(-1)
    uniq_vals, first_pos = np.unique(vals, return_index=True)
    rows = first_pos // aatDiffMat.shape[1] #(整除) 计算出该元素在第几行
    cols = first_pos %  aatDiffMat.shape[1] #(取余) 计算出该元素在第几列
    # 组合坐标并生成字典
    uniq_idx = np.stack([rows, cols], axis=1)
    val2idx = {float(v): (int(r), int(c)) for v, (r,c) in zip(uniq_vals, uniq_idx)}
    return val2idx, uniq_vals, uniq_idx

#模拟物理器件在实际工作中的“读噪声”（Read Noise）或“器件失配”（Device Variation）
def _perturb_coefficients_from_cvals(cvals_np: np.ndarray,
                                     aatData: np.ndarray,
                                     aatDiffMat: np.ndarray,
                                     CV_amp: float,
                                     tol: float = 1e-6) -> np.ndarray:

    cvals = np.round(cvals_np.astype(np.float64), 5)
    val2idx, uniq_vals, uniq_idx = _build_diff_lookup(aatDiffMat)
    new_vals = np.empty_like(cvals, dtype=np.float64)
    Z = np.random.normal(0.0, CV_amp, size=(cvals.size, 2))
    #寻找最接近的物理映射
    for i, cv in enumerate(cvals):
        idx_pair = None
        if cv in val2idx:
            rr, cc = val2idx[cv]
            idx_pair = (rr, cc)
        else:
            diffs = np.abs(uniq_vals - cv)
            j_min = np.argmin(diffs)
            if diffs[j_min] <= tol:
                rr, cc = uniq_idx[j_min]
                idx_pair = (rr, cc)
            else:
                rr, cc = uniq_idx[j_min]
                idx_pair = (rr, cc)
        rr, cc = idx_pair
        a_val = float(aatData[rr])
        b_val = float(aatData[cc])

        #注入差分对乘性噪声
        a_pertb = a_val + a_val * Z[i, 0]
        b_pertb = b_val + b_val * Z[i, 1]
        new_vals[i] = a_pertb - b_pertb

    return new_vals

# ----------------- 全局设备参数定义 -----------------
CV_amp = 0.01    # 代表 5% 的变异系数 (Coefficient of Variation)
sigma1 = 0.5     # 论文中高斯核的宽度参数
sigma2 = 0.5     # 论文中高斯核的宽度参数
CV_sigma1 = 0.01 # sigma 的变异系数
CV_sigma2 = 0.01 # sigma 的变异系数
k = 1.0          # 缩放因子
q = 1.0          # 比例因子
tol = 1e-6
# ---------------------------------------------------

def changeAAT(model, spline_coef):
    spline_coef = spline_coef.to(next(model.parameters()).device)
    for name, param in model.named_parameters():
        if 'spline_weight' in name:
            tensor = param.data
            diff = torch.abs(tensor.reshape(-1, 1) - spline_coef.reshape(1, -1))
            index = torch.argmin(diff, dim=1)
            new_tensor = spline_coef[index].reshape(tensor.shape)
            new_np = new_tensor.detach().cpu().numpy().reshape(-1)
            new_np_pertb = _perturb_coefficients_from_cvals(
                new_np, aatData=aatData, aatDiffMat=aatDiffMat, CV_amp=CV_amp, tol=tol
            )
            new_np_pertb = new_np_pertb.reshape(new_tensor.shape)
            new_tensor_pertb = torch.as_tensor(new_np_pertb, device=tensor.device, dtype=tensor.dtype)
            param.data.copy_(new_tensor_pertb)

class KANLinear(torch.nn.Module):
    def __init__(
        self,
        in_features,
        out_features,
        grid_size=4, # User-defined topological hyperparameter
        spline_order=3,
        scale_noise=0.1,
        scale_base=1.0,
        scale_spline=1.0,
        enable_standalone_scale_spline=False,
        base_activation=torch.nn.ReLU,
        grid_eps=0.02,
        grid_range=[-1, 1], # Default
        spline_weight_init_scale=0.1,
        sigma_left=sigma1 / k,   # 修改：使用实际变量计算
        sigma_right=sigma2 / k   # 修改：使用实际变量计算
    ):
        super(KANLinear, self).__init__()
        self.in_features = in_features
        self.out_features = out_features
        self.grid_size = grid_size
        self.spline_weight_init_scale = spline_weight_init_scale
        grid = torch.linspace(grid_range[0], grid_range[1], grid_size)
        self.register_buffer("grid", grid)
        self.base_weight = torch.nn.Parameter(torch.Tensor(out_features, in_features))
        self.spline_weight = torch.nn.Parameter(
            torch.Tensor(out_features, in_features, grid_size)
        )
        self.sigma_left = sigma_left
        self.sigma_right = sigma_right
        self.gamma = (grid_range[1] - grid_range[0]) / (grid_size - 1)
        self.a = torch.pi * self.gamma
        self.tolerance = 1e-6

        if enable_standalone_scale_spline:
            self.spline_scaler = torch.nn.Parameter(
                torch.Tensor(out_features, in_features)
            )
        self.scale_noise = scale_noise
        self.scale_base = scale_base
        self.scale_spline = scale_spline
        self.enable_standalone_scale_spline = enable_standalone_scale_spline
        self.base_activation = base_activation()
        self.grid_eps = grid_eps
        self.reset_parameters()

    def reset_parameters(self):
        torch.nn.init.kaiming_uniform_(self.base_weight, a=math.sqrt(5) * self.scale_base)
        torch.nn.init.trunc_normal_(self.spline_weight, mean=0, std=self.spline_weight_init_scale)

    def PrewiseRadialBasisFunction_with_noise(self, x: torch.Tensor):
        assert x.dim() == 2 and x.size(1) == self.in_features
        grid: torch.Tensor = self.grid
        bases = torch.zeros(x.size(0), self.in_features, grid.size(0), device=x.device)

        # 修改：替换原本的字符串为实际数值计算
        CV_sigma_left = CV_sigma1 * q
        CV_sigma_right = CV_sigma2 * q

        eps = 1e-12 # Avoid division by zero error
        V_mu = 0.55
        for i in range(grid.size(0)):
            center = grid[i]
            sigma = torch.where(x <= center, torch.tensor(self.sigma_left, device=x.device), torch.tensor(self.sigma_right, device=x.device))
            noise_std_sigma = torch.where(
                x <= center,
                torch.tensor(CV_sigma_left * self.sigma_left, device=x.device),
                torch.tensor(CV_sigma_right * self.sigma_right, device=x.device)
            )
            noise_sigma = torch.normal(mean=0.0, std=noise_std_sigma)
            sigma = sigma + noise_sigma + 1e-6  # 加噪后的sigma
            base_val = torch.exp(-((x - center) ** 2) / (sigma ** 2))
            bases[:, :, i] = base_val
        assert bases.size() == (
            x.size(0),
            self.in_features,
            self.grid_size,
        )
        return bases.contiguous()

    def curve2coeff(self, x: torch.Tensor, y: torch.Tensor):
            assert x.dim() == 2 and x.size(1) == self.in_features
            assert y.size() == (x.size(0), self.in_features, self.out_features)
            A = self.PrewiseRadialBasisFunction(x).transpose(0, 1)
            B = y.transpose(0, 1)
            solution = torch.linalg.lstsq(A, B).solution
            result = solution.permute(2, 0, 1)
            assert result.size() == (
                self.out_features,
                self.in_features,
                self.grid_size,
            )
            return result.contiguous()

    @property
    def scaled_spline_weight(self):
        return self.spline_weight * (
            self.spline_scaler.unsqueeze(-1)
            if self.enable_standalone_scale_spline
            else 1.0
        )

    def scaled_spline_weight_clamped(self):
        return torch.clamp(self.scaled_spline_weight, min=-1, max=1)

    def forward(self, x: torch.Tensor):
        assert x.dim() == 2 and x.size(1) == self.in_features
        base_output = F.linear(self.base_activation(x), self.base_weight)
        spline_output = F.linear(
            self.PrewiseRadialBasisFunction_with_noise(x).view(x.size(0), -1),
            self.scaled_spline_weight_clamped().view(self.out_features, -1),
        )
        return base_output + spline_output  # After the setup, the network includes residual connections based on the memristor differential pairs.

class KAN(torch.nn.Module):
    def __init__(
        self,
        layers_hidden,
        grid_size=4, # User-defined topological hyperparameter
        spline_order=3,
        scale_noise=0.1,
        scale_base=1.0,
        scale_spline=1.0,
        base_activation=torch.nn.ReLU,
        grid_eps=0.02,
        grid_range=[-1, 1], # Default
    ):
        super(KAN, self).__init__()
        self.grid_size = grid_size
        self.spline_order = spline_order
        self.layers = torch.nn.ModuleList()
        for in_features, out_features in zip(layers_hidden, layers_hidden[1:]):
            self.layers.append(
                KANLinear(
                    in_features,
                    out_features,
                    grid_size=grid_size,
                    spline_order=spline_order,
                    scale_noise=scale_noise,
                    scale_base=scale_base,
                    scale_spline=scale_spline,
                    base_activation=base_activation,
                    grid_eps=grid_eps,
                    grid_range=grid_range,
                )
            )
    def forward(self, x: torch.Tensor, update_grid=False):
        for layer in self.layers:
            if update_grid:
                layer.update_grid(x)
            x = layer(x)
        return x

def export_spline_weights(model, filename, sheet_name):
    spline_weights = model.layers[-1].spline_weight.detach().cpu().numpy()
    reshaped_weights = spline_weights.reshape(spline_weights.shape[2], -1)
    df = pd.DataFrame(reshaped_weights)
    if os.path.exists(filename):
        book = load_workbook(filename)
        writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='replace')
        writer.book = book
    else:
        writer = pd.ExcelWriter(filename, engine='openpyxl')
    df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
    writer.close()

transform = transforms.Compose(
    [transforms.Resize(28), transforms.ToTensor(), transforms.Normalize((0.5,), (0.5,))]
)
trainset = torchvision.datasets.MNIST(
    root="./data", train=True, download=True, transform=transform
)
valset = torchvision.datasets.MNIST(
    root="./data", train=False, download=True, transform=transform
)
trainloader = DataLoader(trainset, batch_size=64, shuffle=True)
valloader = DataLoader(valset, batch_size=64, shuffle=False)

model = KAN([28 * 28, 100, 10]) # Define network size
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
model.to(device)
print(f"Number of parameters: {sum(p.numel() for p in model.parameters())}")
#optimizer = optim.AdamW(model.parameters(), lr=0.005, weight_decay=1e-4)

# 在定义 optimizer 之后添加
optimizer = optim.AdamW(model.parameters(), lr=0.005, weight_decay=1e-4)
# 使用余弦退火或步骤衰减
scheduler = optim.lr_scheduler.CosineAnnealingLR(optimizer, T_max=50)
criterion = nn.CrossEntropyLoss()

train_losses = []
train_accuracies = []
val_losses = []
val_accuracies = []
epoch_times = []
iter_accuracies = []

# Initialize the arrays
changeMEM(model, synapse)
changeAAT(model, spline_coef)

for epoch in range(50):
    start_time = time.time()
    model.train()
    train_loss = 0
    train_accuracy = 0
    with tqdm(trainloader) as pbar:
        for i, (images, labels) in enumerate(pbar):
            images = images.view(-1, 28 * 28).to(device)
            optimizer.zero_grad()
            output = model(images)
            loss = criterion(output, labels.to(device))
            loss.backward()
            optimizer.step()
            # 每个 epoch 结束时更新学习率
            scheduler.step()
            changeMEM(model, synapse)
            changeAAT(model, spline_coef)
            train_loss += loss.item()
            accuracy = (output.argmax(dim=1) == labels.to(device)).float().mean().item()
            train_accuracy += accuracy
            pbar.set_postfix(loss=loss.item(), accuracy=accuracy, lr=optimizer.param_groups[0]['lr'])

    train_loss /= len(trainloader)
    train_accuracy /= len(trainloader)
    train_losses.append(train_loss)
    train_accuracies.append(train_accuracy)

    model.eval()
    val_loss = 0
    val_accuracy = 0
    with torch.no_grad():
        for images, labels in valloader:
            images = images.view(-1, 28 * 28).to(device)
            output = model(images)
            val_loss += criterion(output, labels.to(device)).item()
            val_accuracy += (output.argmax(dim=1) == labels.to(device)).float().mean().item()

    val_loss /= len(valloader)
    val_accuracy /= len(valloader)
    val_losses.append(val_loss)
    val_accuracies.append(val_accuracy)
    end_time = time.time()
    epoch_time = end_time - start_time
    epoch_times.append(epoch_time)
    print(f"Epoch {epoch + 1}, Val Loss: {val_loss}, Val Accuracy: {val_accuracy}")

def save_to_excel(data, filename, sheet_name, startcol):
    if os.path.exists(filename):
        try:
            book = load_workbook(filename)
            writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='overlay')
        except Exception as e:
            print(f"Failed to load the Excel file: {e}")
            return
    else:
        writer = pd.ExcelWriter(filename, engine='openpyxl')
    data.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startcol=startcol)
    writer.close()

df_accuracy = pd.DataFrame(train_accuracies, columns=['Accuracy'])
save_to_excel(df_accuracy, './ACC.xlsx', sheet_name='PR_MNIST', startcol=0)

print("Completed")